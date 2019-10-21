import os
import sqlite3
import email
import smtplib
import imaplib
import json
import openpyxl
import contextlib
import sys
import csv

from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from email.mime.base import MIMEBase

with open("config.json","r") as config_file:
    cfg = json.load(config_file)

def check_for_lab_email(mail):  # CHECK FOR SPECIFIC EMAIL (From and Unopened)
    success = 0
    filename = ' '
    result, msgs = mail.search (None,
                                'UNSEEN',
                                'FROM', cfg["imapfrom"])
    msgs = msgs[0].split ()
    for emailid in msgs:
        result, data = mail.fetch (emailid, '(RFC822)')
        email_body = data[0][1]
        m = email.message_from_bytes (email_body)
        if m.get_content_maintype () != 'multipart':
            continue
        for part in m.walk ():
            if part.get_content_maintype () == 'multipart':
                continue
            if part.get ('Content-Disposition') is None:
                continue
            filename = part.get_filename ()
            if cfg['attach_keyword'] in filename and 'xlsx' in filename:
                print ("New EMAIL w/attachment Found")
                sv_path = os.path.join (cfg["data_dir"], filename)
                fp = open(filename, 'wb')
                fp.write (part.get_payload (decode=True))
                fp.close ()
                print ('Spreadsheet retrieved and saved --->  ', filename)
                success = 1
    mail.logout ()
    return success, filename

def cleanup():
    filelist = [os.path.join (cfg["data_dir"], "Samples.csv"),
                os.path.join (cfg["data_dir"], "rawdata.xlsx")]
    with contextlib.suppress (FileNotFoundError):
        for f in filelist:
            os.remove(f)
    print("Cleanup completed")

def login_mail():
    try:
        mail = imaplib.IMAP4_SSL(cfg["imapserver"])
        mail.login(cfg["imaplogon"],cfg["imappwd"])
        mail.select('inbox')
        return mail

    except:  
        print("Unable to login to mail server...aborting")
        sys.exit(1)  # replace with retry someday!                              

def get_attach(msgs,mail):    
        msgs = msgs[0].split()                        
        for emailid in msgs:
            result, data = mail.fetch(emailid, '(RFC822)')
            email_body = data[0][1]
            m = email.message_from_bytes(email_body)
            if m.get_content_maintype() != 'multipart':
                continue
            for part in m.walk():            
                if part.get_content_maintype() == 'multipart':
                    continue
                if part.get('Content-Disposition') is None:
                    continue
                filename=part.get_filename()         
                if cfg['attach_keyword'] in filename and 'xlsx' in filename:
                    print("New EMAIL w/attachment Found")
                    sv_path = os.path.join(cfg["local_save_dir"], filename)
                    fp = open(sv_path, 'wb')
                    fp.write(part.get_payload(decode=True))
                    fp.close()
                    print('Spreadsheet retrieved and saved')

def insert_data(filename):
    try:                ##Open the DB
        dbName = os.path.join (cfg["data_dir"], "GusHealth.db")
        con = sqlite3.connect(dbName)
        c = con.cursor()
    except:
        print("ERROR:  Unable to connect to DB")

    try:
        wb = openpyxl.load_workbook (os.path.join(cfg["data_dir"],filename), data_only=True)
        ws=wb.active
    except:
        print('workbook not opened')

# Find the last row with sample data
    for last_row in range(3,100):
        if ws.cell(last_row,1).value == None:
            last_row = last_row - 1
            num_rows = last_row - 2
            break

# Find the last column with a DATE
    for last_col in range(3, 35):
        if ws.cell(1,last_col).value == None:
            last_col = last_col -1
            col_letter = get_column_letter(last_col)
            break

#Create lists based on the row/col from above
    account = []
    for x in range (3, last_row+1):
        account.append (ws.cell (x, 1).value)
    doc = []
    for x in range (3, last_row+1):
        doc.append (ws.cell (x, 2).value)

# Insert records into DB only if sample >0
    insert_count = 0
    for y in range (4, last_col+1):
        sample_date = ws.cell (1, y).value.strftime ('%Y-%m-%d')
        samples = []
        for x in range (3, last_row+1):
            samples.append (ws.cell (x, y).value)
        for x in range (0,num_rows):  ## Insert rows into DB
            if samples[x] is not None and samples[x] > 0:
                c.execute ("""INSERT or REPLACE INTO Samples VALUES(?,?,?,?)""",
                          (sample_date, account[x], doc[x], samples[x]))
                insert_count=insert_count+1
        con.commit()
    con.close()
    print("Database updated")
    print (insert_count, " records updated")
    query_date = sample_date[0:7]
    return query_date


def create_csv_file(sample_query):
    try:            ##Open the DB
        dbName = os.path.join (cfg["data_dir"], "GusHealth.db")
        con = sqlite3.connect (dbName)
        c = con.cursor ()
    except:
        print ("ERROR:  Unable to connect to DB")

    sample_query = sample_query + "-%"
    c.execute ('SELECT * FROM ClicData WHERE date like ?', (sample_query,))
    data = c.fetchall()

    Samples = os.path.join (cfg["data_dir"], "Samples.csv")
    with open(Samples,'w',newline='') as f_handle:
        writer = csv.writer(f_handle)
        header = ['Date','Practice','Doctor','Samples']
        writer.writerow(header)
        for row in data:
            writer.writerow(row)
        print("Query created")

def send_mail():   
    smtp_server =       cfg['smtpserver']
    smtp_port =         cfg['smtpport']
    smtp_login =        cfg['smtplogon']
    smtp_password =     cfg['smtppwd']
    to_addr =           cfg['recipient1']
    cc_addr =           " "
    group =             [to_addr, cc_addr]
    body =              cfg['body1']

    msg = MIMEMultipart()
    msg['From'] =       cfg['from']
    msg['To'] =         to_addr + ',' + cc_addr
    msg['Subject'] =    cfg['subject1']
    msg.attach(MIMEText(body, 'plain'))

    text = msg.as_string()
    try: #  INITIALIZE AND LOGON TO SMTP SERVER
        smtp_server = smtplib.SMTP(smtp_server, smtp_port)
        smtp_server.ehlo()  # Send mandatory 'hello' msg
        smtp_server.starttls() # Start TLS Encryption as we're not using SSL.
        smtp_server.login(smtp_login,smtp_password)# login
        smtp_server.sendmail(from_addr, group, text)   # SEND THE EMAIL
        smtp_server.quit()
        print("Email(s) sent")
        print("Email logout")
    except Exception:
        print("\n*** Email FAILED ***")

def send_mail_attachment():
    smtp_server = cfg['smtpserver2']
    smtp_port = cfg['smtpport2']
    smtp_login = cfg['smtplogon2']
    smtp_password = cfg['smtppwd2']
    from_addr = cfg['from2']
    to_addr = cfg['recipient2']
    cc_addr = " "
    group = [to_addr, cc_addr]
    body = cfg['body2']

    msg = MIMEMultipart()
    msg['From'] = cfg['from2']
    msg['To'] = to_addr + ',' + cc_addr
    msg['Subject'] = cfg['subject2']
    msg.attach(MIMEText(body, 'plain'))

    filename = cfg['data_dir'] + '/' + cfg['query_file']
    attachment = open(filename, "rb")
    p = MIMEBase('application', 'octet-stream')
    p.set_payload((attachment).read())
    encoders.encode_base64(p)
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
    msg.attach(p)
    text = msg.as_string()

    try:  # INITIALIZE AND LOGON TO SMTP SERVER
        smtp_server = smtplib.SMTP(smtp_server, smtp_port)
        smtp_server.ehlo()  # Send mandatory 'hello' msg
        smtp_server.starttls()  # Start TLS Encryption as we're not using SSL.
        smtp_server.login(smtp_login, smtp_password)  # login
        smtp_server.sendmail(from_addr, group, text)  # SEND THE EMAIL
        smtp_server.quit()
        print("File sent to ClicData")
    except Exception:
        print("\n*** Email FAILED ***")

def copy_to_google_drive():
    # COPY FILE (rawdata.csv) TO GOOGLE DRIVE
    sv_path = os.path.join(cfg["local_save_dir"], "Samples.csv")
    try:
        subprocess.call(['rclone', 'copy', sv_path, 'googledrive:'], -1)
        print("File (samples.csv) copied to Google Drive")
    except(exception):
        print("*** ERROR COPYING FILE TO GOOGLE DRIVE")
        print(exception)

def logit(log_data):
    print(log_data[0],log_data[1],log_data[2],log_data[3]) # write to console
    loginfo = {
            "Date":     log_data[0],
            "Time":     log_data[1],
            "Event":    log_data[2],
            "Duration": log_data[3]
              }                  
    with open("/var/www/html/log.json","w+") as log_file:
        log_file.write("[")
        logrecord = ("\n" + json.dumps(loginfo) + "]")
        log_file.write(logrecord)
