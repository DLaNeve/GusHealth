import os
import sqlite3
import email
import smtplib
import imaplib
import json
import openpyxl
import datetime
import time

from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


  # OPEN/LOAD/CLOSE CONFIG FILE
with open("config.json","r") as config_file:
    cfg = json.load(config_file)

def cleanup():
    try:
        os.remove(cfg["local_save_dir"] + "/rawdata.xlsx") 
        os.remove(cfg["local_save_dir"] + "/Samples.csv")
        print("Cleanup completed\n")
    except:
        print("No cleanup necessary")
    
def login_mail(mail):
    try:
        mail = imaplib.IMAP4_SSL(cfg["imapserver"])
        mail.login(cfg["imaplogon"],cfg["imappwd"])
        mail.select('inbox')
        print("Email login")
    except:  
        print("Unable to login to mail server...aborting")
        sys.exit(1)  # replace with retry someday!                              

def get_attach(msgs,mail):    
        msgs = msgs[0].split()                        
        for emailid in msgs:
            result, data = mail.fetch(emailid, '(RFC822)')
            email_body = data[0][1]
            m = email.message_from_bytes(email_body)
            if m.get_content_maintype() != 'multipart':
                continue
            for part in m.walk():            
                if part.get_content_maintype() == 'multipart':
                    continue
                if part.get('Content-Disposition') is None:
                    continue
                filename=part.get_filename()         
                if cfg['attach_keyword'] in filename and 'xlsx' in filename:
                    print("New EMAIL w/attachment Found")
                    sv_path = os.path.join(cfg["local_save_dir"], filename)
                    fp = open(sv_path, 'wb')
                    fp.write(part.get_payload(decode=True))
                    fp.close()
                    print('Spreadsheet retrieved and saved')
                    print(sv_path)



                    
def insert_data(sv_path):
    ##Open the DB
    try:
        con = sqlite3.connect(cfg["local_save_dir"] + "/GusHealth.db")
        c = con.cursor()
    except:
        print("ERROR:  Unable to connect to DB")
        
    ##Open the workbook
    wb=openpyxl.load_workbook(sv_path,data_only=True)
    ws=wb.active    

    ##Find the row containing the word TOTAL.
    col_num = 2
    for total_row in range(1,70):
        if ws.cell(total_row,col_num).value == "TOTAL":
            print("The line containing the TOTALs is: ",total_row)
            break
        
    ##Find the last column with total data
    for max_col_num in range(1, 35):      
        if ws.cell(total_row,max_col_num).value == 0:
            max_col_num = max_col_num-1
            col_letter = get_column_letter(max_col_num)
            print("The Column containing the current data is: ", col_letter)
            break      

    ##Find the last row containing data.  Assume the first row is 3
    col_num = 1
    for max_row_num in range(3,total_row):
        if ws.cell(max_row_num,col_num).value is None:
            break  
    total_rows = max_row_num - 3
    print("The last row containing data is: ",max_row_num-1)

        
    ##Create lists based on the row/col from above

    account=[]    
    for x in range(3,max_row_num):
        account.append(ws.cell(x,1).value)
        
    doc = []
    for x in range(3,max_row_num):
        doc.append(ws.cell(x,2).value)

    samples = []
    for x in range(3,max_row_num):
        samples.append(ws.cell(x,max_col_num).value)

    format = '%Y-%m-%d'
    sample_date = ws.cell(1,max_col_num).value.strftime('%Y-%m-%d')  
    epoch_date = time.mktime(time.strptime(sample_date, format))
    epoch_date = int(epoch_date)
    print("The date for the imported samples is: ",sample_date)
    print("The Epoch for the imported samples is: ",epoch_date)
    
    ## Insert rows into DB 
    for y in range(0,total_rows):
        try:
            if samples[y] > 0:
                c.execute("""INSERT INTO Samples VALUES(?,?,?,?,?)""",(sample_date,account[y],doc[y],samples[y],epoch_date))                
        except:
            print("ERROR:  Unable to Insert row ",y," into Samples Table")




    try:      
        con.commit()
        con.close()
    except:
        print("ERROR: Commit failed")

def send_mail():   
    smtp_server =       cfg['smtpserver']
    smtp_port =         cfg['smtpport']
    smtp_login =        cfg['smtplogon']
    smtp_password =     cfg['smtppwd']
    to_addr =           cfg['recipient1']
    cc_addr =           cfg['recipient2']
    group =             [to_addr, cc_addr]
    from_addr =         cfg['from']
    subject_line =      cfg['subject']
    body_msg =          cfg['body1']  
    msg =               MIMEMultipart()
    msg['From'] =       from_addr
    msg['To'] =         to_addr + ',' + cc_addr
    msg['Subject'] =    subject_line 
    body =              body_msg
           
    msg.attach(MIMEText(body, 'plain'))
    text = msg.as_string()
    try: #  INITIALIZE AND LOGON TO SMTP SERVER
        smtp_server = smtplib.SMTP(smtp_server, smtp_port) # Specify Gmail Mail server
        smtp_server.ehlo()  # Send mandatory 'hello' message to SMTP server
        smtp_server.starttls() # Start TLS Encryption as we're not using SSL.
        smtp_server.login(smtp_login,smtp_password)# login
        smtp_server.sendmail(from_addr, group, text)   # SEND THE EMAIL
        smtp_server.quit()
        print("Email(s) sent")
        print("Email logout")
    except Exception:
        print("\n*** Email FAILED ***")


def logit(log_data):
    print(log_data[0],log_data[1],log_data[2],log_data[3]) # write to console
    loginfo = {
            "Date":     log_data[0],
            "Time":     log_data[1],
            "Event":    log_data[2],
            "Duration": log_data[3]
              }                  
    with open("/var/www/html/log.json","w+") as log_file:
        log_file.write("[")
        logrecord = ("\n" + json.dumps(loginfo) + "]")
        log_file.write(logrecord)    
            
