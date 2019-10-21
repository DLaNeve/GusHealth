import sqlite3
import json
import openpyxl
from openpyxl.utils import get_column_letter


##Open the workbook to view ONLY the DATA not the formulas
wb=openpyxl.load_workbook('/home/pi/Sean/v5/MARIA MARCH 2018.xlsx',data_only=True)
ws=wb.active

##Find the row containing the word TOTAL.
col_num = 2
for total_row in range(1,50):
    if ws.cell(total_row,col_num).value == "TOTAL":
        print("The line containing the TOTALs is: ",total_row)
        break

##Find the last column with total data
for max_col_num in range(1, 35):      
    if ws.cell(total_row,max_col_num).value == 0:
        max_col_num = max_col_num-1
        col_letter = get_column_letter(max_col_num)
        print("The Column containing the current data is: ", col_letter)
        break

##Find the last row containing data.  Assume the first row is 3
col_num = 1
for max_row_num in range(3,total_row):
    if ws.cell(max_row_num,col_num).value is None:
        break
total_rows = max_row_num - 3

print("The last row containing data is: ",max_row_num)
print("the total number of records are: ",total_rows)
        

##Create lists based on the row/col from above
sample_date = ws.cell(1,max_col_num).value.strftime('%m-%d-%y')
print("The date for the imported samples is: ",sample_date)

account=[]    
for x in range(3,max_row_num):
    account.append(ws.cell(x,1).value)
    
doc = []
for x in range(3,max_row_num):
    doc.append(ws.cell(x,2).value)

samples = []
for x in range(3,max_row_num):
    samples.append(ws.cell(x,max_col_num).value) 

##Database stuff
##try:
##    con = sqlite3.connect('/home/pi/Sean/v5/GusHealth.db')    
##    c = con.cursor()
##except:
##    print("ERROR:  Unable to connect to DB")
##   
##for y in range(0,total_rows):
##    try:
##        if samples[y] > 0:
##            c.execute("""INSERT INTO Samples VALUES(?,?,?,?)""",(sample_date,account[y],doc[y],samples[y]))
##    except:
##        print("ERROR:  Unable to Insert row ",y," into Samples Table")
##try:    
##    con.commit()
##    con.close()
##except:
##    print("ERROR: Commit failed")

