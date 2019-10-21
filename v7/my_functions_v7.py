import os
import sqlite3
import email
import smtplib
import imaplib
import json
import openpyxl
import contextlib
import csv
import time

from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from email.mime.base import MIMEBase

with open("config.json","r") as config_file:
    cfg = json.load(config_file)

# def check_for_lab_email(mail):  # CHECK FOR SPECIFIC EMAIL (From and Unopened)
#     success = 0
#     filename = ' '
#     attach = ' '
#     result, msgs = mail.search (None,
#                                 'UNSEEN',
#                                 'FROM', cfg["imapfrom"])
#     msgs = msgs[0].split ()
#     for emailid in msgs:
#         result, data = mail.fetch (emailid, '(RFC822)')
#         email_body = data[0][1]
#         m = email.message_from_bytes (email_body)
#         if m.get_content_maintype () != 'multipart':
#             continue
#         for part in m.walk ():
#             if part.get_content_maintype () == 'multipart':
#                 continue
#             if part.get ('Content-Disposition') is None:
#                 continue
#             attach = part.get_filename ()
#             if cfg['attach_keyword'] in attach and 'xlsx' in attach:
#                 filename = attach
#                 print ("New EMAIL w/attachment Found")
#                 fp = open(filename, 'wb')
#                 fp.write (part.get_payload (decode=True))
#                 fp.close ()
#                 print ('Spreadsheet retrieved and saved --->  ', filename)
#                 success = 1
#     return success, filename

# def cleanup():
#     filelist = ["Samples.csv",
#                 "rawdata.xlsx"]
#     with contextlib.suppress (FileNotFoundError):
#         for f in filelist:
#             os.remove(f)
#     print("Cleanup completed")

# def login_mail():
#     while True:
#         try:
#             mail = imaplib.IMAP4_SSL(cfg["imapserver"])
#             mail.login(cfg["imaplogon"],cfg["imappwd"])
#             mail.select('inbox')
#             return mail
#         except:
#             print("Unable to login to mail server...retrying")
#             time.sleep (10)
#             continue
#         break

# def insert_clic_data(filename):
#     try:                ##Open the DB
#         dbName =  cfg["db_name"]
#         con = sqlite3.connect(dbName)
#         c = con.cursor()
#     except:
#         print("ERROR:  Unable to connect to DB")
#     try:
#         wb = openpyxl.load_workbook (filename, data_only=True)
#         ws=wb.active
#     except:
#         print('workbook not opened')
#
# # Find the last row with sample data
#     for last_row in range(3,100):
#         if ws.cell(last_row,1).value == None:
#             last_row = last_row - 1
#             num_rows = last_row - 2
#             break
# # Find the last column with a DATE
#     for last_col in range(3, 35):
#         if ws.cell(1,last_col).value == None:
#             last_col = last_col -1
#             col_letter = get_column_letter(last_col)
#             break
# #Create lists based on the row/col from above
#     account = []
#     for x in range (3, last_row+1):
#         account.append (ws.cell (x, 1).value)
#     doc = []
#     for x in range (3, last_row+1):
#         doc.append (ws.cell (x, 2).value)
# # Insert records into DB only if sample >0
#     insert_count = 0
#     for y in range (4, last_col+1):
#         sample_date = ws.cell (1, y).value.strftime ('%Y-%m-%d')
#         samples = []
#         for x in range (3, last_row+1):
#             samples.append (ws.cell (x, y).value)
#         for x in range (0,num_rows):  ## Insert rows into DB
#             if samples[x] is not None and samples[x] > 0:
#                 c.execute("""INSERT or REPLACE INTO Samples VALUES(?,?,?,?,?,?,?)""",
#                           (" ",sample_date, account[x], doc[x], samples[x],0,0),)
#                 insert_count=insert_count+1
#         con.commit()
#     con.close()
#     print("Database updated")
#     print (insert_count, " records updated")
#     query_date = sample_date[0:7]
#     return query_date

# def create_clic_file():
#     sample_query = sample_query + "-%"
#     c.execute ('SELECT Date, Practice, Doctor, Samples FROM Samples WHERE date like ?', (sample_query,))
#     data = c.fetchall()
#
#     Samples = "Samples.csv"
#     with open(Samples,'w',newline='') as f_handle:
#         writer = csv.writer(f_handle)
#         header = ['Date','Practice','Doctor','Samples']
#         writer.writerow(header)
#         for row in data:
#             writer.writerow(row)
#         print("Query created")

# def send_mail_attachment():
#     smtp_server = cfg['smtpserver2']
#     smtp_port = cfg['smtpport2']
#     smtp_login = cfg['smtplogon2']
#     smtp_password = cfg['smtppwd2']
#     from_addr = cfg['from2']
#     to_addr = cfg['recipient2']
#     cc_addr = " "
#     group = [to_addr, cc_addr]
#     body = cfg['body2']
#
#     msg = MIMEMultipart()
#     msg['From'] = cfg['from2']
#     msg['To'] = to_addr + ',' + cc_addr
#     msg['Subject'] = cfg['subject2']
#     msg.attach(MIMEText(body, 'plain'))
#
#     filename = cfg['query_file']
#     attachment = open(filename, "rb")
#     p = MIMEBase('application', 'octet-stream')
#     p.set_payload((attachment).read())
#     encoders.encode_base64(p)
#     p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
#     msg.attach(p)
#     text = msg.as_string()
#
#     try:  # INITIALIZE AND LOGON TO SMTP SERVER
#         smtp_server = smtplib.SMTP(smtp_server, smtp_port)
#         smtp_server.ehlo()  # Send mandatory 'hello' msg
#         smtp_server.starttls()  # Start TLS Encryption as we're not using SSL.
#         smtp_server.login(smtp_login, smtp_password)  # login
#         smtp_server.sendmail(from_addr, group, text)  # SEND THE EMAIL
#         smtp_server.quit()
#         print("File sent to ClicData")
#     except Exception:
#         print("\n*** Email FAILED ***")

# def send_text():
#     smtpserver = cfg['smtpserver']
#     smtpport = cfg['smtpport']
#     gmailaddress = cfg['smtplogon']
#     gmailpassword = cfg['smtppwd']
#     mailto1 = cfg['recipient1']
#     mailto2 = cfg['recipient2']
#     group = [mailto1,mailto2]
#     msg = cfg['msg']
#     try: #  INITIALIZE AND LOGON TO SMTP SERVER
#         mailServer = smtplib.SMTP(smtpserver , smtpport)
#         mailServer.starttls()
#         mailServer.login(gmailaddress , gmailpassword)
#         mailServer.sendmail(gmailaddress, group , msg)
#         mailServer.quit()
#         print("Text(s) sent")
#     except Exception:
#         print("\n*** Text FAILED ***")