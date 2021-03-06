#!/usr/bin/python3
# version 10-Aug

import time
from time import strptime
import datetime
import os
import re
import calendar
import json
import contextlib
import email
import smtplib
import imaplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from email.mime.base import MIMEBase
from tkinter import simpledialog
from tkinter import filedialog
from tkinter import *


import sqlite3

from openpyxl.utils import get_column_letter
import openpyxl

import csv
import threading

with open("config.json","r") as config_file:
    cfg = json.load(config_file)
root = Tk()

def login_mail():
    while True:
        try:
            mail = imaplib.IMAP4_SSL(cfg["imapserver"])
            mail.login(cfg["imaplogon"],cfg["imappwd"])
            mail.select('inbox')
            return mail
        except:
            msg = "Unable to login to mail server...retrying"
            print(msg)
            T.insert('1.0',msg)
            time.sleep (10)
            continue
        break

def check_for_lab_email(mail):  # CHECK FOR SPECIFIC EMAIL (From and Unopened)
    global success, filename
    success = 0
    filename = ' '
    attach = ' '
    result, msgs = mail.search (None,
                                'UNSEEN',
                                'FROM', cfg["imapfrom"])
    msgs = msgs[0].split ()
    for emailid in msgs:
        result, data = mail.fetch (emailid, '(RFC822)')
        email_body = data[0][1]
        m = email.message_from_bytes (email_body)
        if m.get_content_maintype () != 'multipart':
            continue
        for part in m.walk ():
            if part.get_content_maintype () == 'multipart':
                continue
            if part.get ('Content-Disposition') is None:
                continue
            attach = part.get_filename ()
            if cfg['attach_keyword'] in attach and 'xlsx' in attach:
                filename = attach
                msg = "New EMAIL w/attachment Found\n"
                print (msg)
                T.insert('1.0',msg)
                fp = open(filename, 'wb')
                fp.write (part.get_payload (decode=True))
                fp.close ()
                print ('Spreadsheet retrieved and saved --->  ', filename)
                msg = 'Spreadsheet retrieved-->  ' + filename + "\n"
                T.insert('1.0',msg)
                success = 1
    return success, filename

def check_for_override_email(mail):
    success = 0
    filename = ' '
    attach = ' '
    result, msgs = mail.search (None,
                                'UNSEEN',
                                'SUBJECT', "OVERRIDE")
    msgs = msgs[0].split ()
    for emailid in msgs:
        result, data = mail.fetch (emailid, '(RFC822)')
        email_body = data[0][1]
        m = email.message_from_bytes (email_body)
        if m.get_content_maintype () != 'multipart':
            continue
        for part in m.walk ():
            if part.get_content_maintype () == 'multipart':
                continue
            if part.get ('Content-Disposition') is None:
                continue
            attach = part.get_filename ()
            if cfg['attach_keyword'] in attach and 'xlsx' in attach:
                filename = attach
                print ("New EMAIL w/attachment Found")
                fp = open (filename, 'wb')
                fp.write (part.get_payload (decode=True))
                fp.close ()
                print ('Spreadsheet retrieved and saved --->  ', filename)
                msg = 'Spreadsheet retrieved and saved --->' + filename
                T.insert('1.0',msg)
                success = 1
    return success, filename

def insert_samples():
# Find the first row with sample data
    for first_row in range(2,100):
        if ws.cell(first_row,2).value != None:
            break
# Find the last row with sample data
    for last_row in range(first_row,100):
        if ws.cell(last_row,2).value == None:
            last_row = last_row - 1
            num_rows = last_row - 2
            break
# Find the last column with a DATE
    for last_col in range(first_row, 35):
        if ws.cell(1,last_col).value == None:
            last_col = last_col -1
            col_letter = get_column_letter(last_col)
            break
#Create lists based on the row/col from above
    account = []
    for x in range (first_row, last_row+1):
        if ws.cell (x, 1).value is None:
            ws.cell (x,1).value = ws.cell (x-1, 1).value
        account.append (ws.cell (x, 1).value)
    doc = []
    for x in range (first_row, last_row+1):
        doc.append (ws.cell (x, 2).value)
# Insert records into DB only if sample >0
    insert_count = 0
    total_samples = 0
    for y in range (4, last_col+1):
        global sample_date, my_month
        sample_date = ws.cell (1, y).value
        sample_date = sample_date.strftime ('%Y-%m-%d')
        my_month = calendar.month_name[int(sample_date[5:7])]
        samples = []
        for x in range (first_row, last_row+1):
            samples.append (ws.cell (x, y).value)
        for x in range (0,num_rows):  ## Insert rows into DB
            if samples[x] is not None and samples[x] > 0:
                c.execute ("""INSERT or REPLACE INTO Samples VALUES(?,?,?,?,?,?,?)""",
                          (my_month, sample_date, account[x], doc[x], samples[x],0,0))
                insert_count=insert_count+1
                total_samples = total_samples + samples[x]
        con.commit()
        c.execute ("UPDATE samples "
                   "SET AcctNo = (SELECT accounts.AcctNo FROM accounts WHERE accounts.ReportName = Samples.Practice),"
                   "RepNo =  (SELECT accounts.RepNo FROM accounts WHERE accounts.ReportName = Samples.Practice)"
                   "WHERE samples.Month =?",(my_month,))
        con.commit()

    c.execute ("select practice from SAMPLEs where acctno is null and month=?",(my_month,))
    problem = c.fetchall ()
    print (dbName, "updated:", "Total Samples :",total_samples, "   Records:", insert_count,"   Errors:", len(problem))
    msg = "\nTotal Samples: " + str(total_samples) + " Records: " + str(insert_count) + " Errors: " + str(len(problem)) + "\n"
    T.insert ('1.0', msg)
    return

def send_clik_mail():
    smtp_server = cfg['smtpserver2']
    smtp_port = cfg['smtpport2']
    smtp_login = cfg['smtplogon2']
    smtp_password = cfg['smtppwd2']
    from_addr = cfg['from2']
    to_addr = cfg['recipient2']
    cc_addr = " "
    group = [to_addr, cc_addr]
    body = cfg['body2']

    msg = MIMEMultipart()
    msg['From'] = cfg['from2']
    msg['To'] = to_addr + ',' + cc_addr
    msg['Subject'] = cfg['subject2']
    msg.attach(MIMEText(body, 'plain'))

    filename = cfg['query_file']
    attachment = open(filename, "rb")
    p = MIMEBase('application', 'octet-stream')
    p.set_payload((attachment).read())
    encoders.encode_base64(p)
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
    msg.attach(p)
    text = msg.as_string()

    try:  # INITIALIZE AND LOGON TO SMTP SERVER
        smtp_server = smtplib.SMTP(smtp_server, smtp_port)
        smtp_server.ehlo()  # Send mandatory 'hello' msg
        smtp_server.starttls()  # Start TLS Encryption as we're not using SSL.
        smtp_server.login(smtp_login, smtp_password)  # login
        smtp_server.sendmail(from_addr, group, text)  # SEND THE EMAIL
        smtp_server.quit()
        msg = "File sent to ClicData\n"
        print(msg)
        T.insert('1.0',msg)

    except Exception:
        print("\n*** Email FAILED ***")
    return

def send_text(num_problems):
    smtpserver = cfg['smtpserver']
    smtpport = cfg['smtpport']
    gmailaddress = cfg['smtplogon']
    gmailpassword = cfg['smtppwd']
    mailto3 = cfg['recipient3']
    mailto4 = cfg['recipient4']
    mailto5 = cfg['recipient5']
    group = [mailto3,mailto4,mailto5]
    msg = cfg['msg'] + "Errors= " + str(num_problems)
    try: #  INITIALIZE AND LOGON TO SMTP SERVER
        mailServer = smtplib.SMTP(smtpserver , smtpport)
        mailServer.starttls()
        mailServer.login(gmailaddress , gmailpassword)
        mailServer.sendmail(gmailaddress, group , msg)
        mailServer.quit()
        msg= "Text(s) sent"
        print(msg)
        T.insert('1.0',msg)
    except Exception:
        print("\n*** Text FAILED ***")

def cleanup():
    filelist = ["Samples.csv",
                "rawdata.xlsx"]
    with contextlib.suppress (FileNotFoundError):
        for f in filelist:
            os.remove(f)
    msg = "Cleanup completed\n"
    print(msg)
    T.insert('1.0',msg)

def process_clicdata():
    global dbName, con, c
    dbName = cfg["db_name"]
    con = sqlite3.connect (dbName,check_same_thread=False)
    c = con.cursor ()
    success = 0
    freq = int (cfg["mail_chk_freq"]) * 60
    while True:
        mail = login_mail ()  # login to GusHealth Email
        normal_mail_status = check_for_lab_email (mail)  # chk for proper email-save the attachment
        over_ride_status = check_for_override_email(mail)  # chk for the word OVERRIDE in the subject line the attachment
        if normal_mail_status[0] == 1:
            success = normal_mail_status[0]
            filename = normal_mail_status[1]
        elif over_ride_status[0] == 1:
            success = over_ride_status[0]
            filename = over_ride_status[1]

        global ws
        try:
            wb = openpyxl.load_workbook (filename, data_only=True)
            ws = wb.active
        except:
            msg = '\nNothing to do...\n'
            print (msg)
            T.insert('1.0', msg)

        if success:
            insert_samples()
            c.execute ("select practice from SAMPLEs where acctno is null and month=?", (my_month,))
            problem = c.fetchall ()
            num_problems = len(problem)
            if problem:
                print ("\n")
                for g in range (0, len (problem)):
                    print ("    *** ERROR: Cannot identify Account for SAMPLE record --> ", problem[g])
                    msg = "SAMPLES-Unknown Practice-->" + str (problem[g]) + "\n"
                    T.insert ('1.0', msg)

            c.execute ('SELECT Date, Practice, Doctor, Samples FROM Samples WHERE month = ?', (my_month,))
            data = c.fetchall ()
            Samples = "Samples.csv"
            with open (Samples, 'w', newline='') as f_handle:
                writer = csv.writer (f_handle)
                header = ['Date', 'Practice', 'Doctor', 'Samples']
                writer.writerow (header)
                for row in data:
                    writer.writerow (row)
            send_clik_mail()
            send_text(num_problems)
            normal_mail_status = 0 * 2
            success = 0
            cleanup()
        else:
            now = datetime.datetime.now ()
            print ("Current date and time: ")
            print (now.strftime ("%Y-%m-%d %H:%M:%S"))
            print ('Sleeping', freq / 60, "mins")
            T.insert('1.0',"Current date and time: " + now.strftime ("%Y-%m-%d %H:%M:%S")+ "\n")
            T.insert('1.0',"Sleeping " + str(freq/60) + " mins\n")
            time.sleep (int (freq))

def start_thread():
    ClicData_Proc=threading.Thread(target=process_clicdata)
    ClicData_Proc.daemon = True
    ClicData_Proc.start()

def get_file():
    def sel():
        global ws, my_month
        my_month = wb.sheetnames[(choice.get())]
        ws = wb[my_month]
        print ("Month Selected --> ", my_month)
        msg = "Month Selected --> " + my_month + "\n"
        T.insert('1.0',msg)
        insert_all_data()

        win2.destroy()
        return

    global fname, wb
    fname = filedialog.askopenfilename \
        (initialdir="/", title="Select file",
         filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
    wb = openpyxl.load_workbook (fname, data_only=True)

    msg = "File Selected -->" + os.path.split(fname)[1] + "\n"
    T.insert ('1.0', msg)

    win2=Toplevel(root)
    win2.geometry ('150x300')
    choice = IntVar ()
    rb = [0] * 12
    for x in range(0,len(wb.sheetnames)):
        rb[x] = Radiobutton(win2, text=wb.sheetnames[x], variable=choice, value=x, command=sel).pack (anchor=W)
    frame.pack ()
    win2.mainloop()
    wb.close()
    return

def insert_monthly_data():

    ##Find the row containing the words "Referring... (block 1)
    col_num = 1
    for block1_start in range (1, 500):
        if ws.cell (block1_start, col_num).value == "Referring Provider Name":
        #if re.match ("Referring", ws.cell (block1_start, col_num).value):
            block1_start = block1_start + 1
            break
    ##Find the last row containing data in block 1.
    col_num = 1
    for block1_stop in range (block1_start, 500):
        if ws.cell (block1_stop, col_num).value is None:
            block1_stop = block1_stop - 1
            break
    ##Find the row containing the word "Referring... again  (block 2).
    col_num = 1
    for block2_start in range (block1_stop, 500):
        if ws.cell (block2_start, col_num).value == "Referring Provider Name":
        #if re.match ("Referring", ws.cell (block2_start, col_num).value):
            block2_start = block2_start + 1
            break
    ##Find the last row containing data in block 2
    col_num = 1
    for block2_stop in range (block2_start, 500):
        if ws.cell (block2_stop, col_num).value is None:
            block2_stop = block2_stop - 1
            break
    # determine which column contains specific data in 1st block
    for i in range(1,15):
        try:
            if re.match("Referring",ws.cell(block1_start-1,i).value):
                provider_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Practice", ws.cell (block1_start-1, i).value):
                practice_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Account", ws.cell (block1_start-1, i).value):
                CaseNo_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Service", ws.cell (block1_start-1, i).value):
                svc_date_col = i
                break
        except:
            continue
    insert_count = 0
    for z in range (block1_start, block1_stop + 1):
        selfpay_date = ws.cell (z,svc_date_col).value
        selfpay_date = selfpay_date.strftime ('%Y-%m-%d')
        practice = ws.cell(z,practice_col).value
        provider = ws.cell(z,provider_col).value
        CaseNo = ws.cell(z,CaseNo_col).value
        c.execute ("""INSERT or REPLACE INTO selfpay VALUES(?,?,?,?,?,?,?)""",
                   (my_month, selfpay_date, practice, provider, CaseNo, 0, 0))
        insert_count = insert_count + 1
    con.commit ()
    # associate each SAMPLE record with the proper Rep
    c.execute("UPDATE selfpay "
              "SET AcctNo = (SELECT accounts.AcctNo FROM accounts "
              "WHERE accounts.MonthlyName = selfpay.Practice),"
              "RepNo =  (SELECT accounts.RepNo FROM accounts WHERE accounts.MonthlyName = selfpay.Practice)"
              "where selfpay.Month=?",(my_month,))
    con.commit()

    c.execute ("select practice from selfpay where acctno is null and month=?", (my_month,))
    problem = c.fetchall ()

    print(dbName, "updated:", "Total SelfPays: ", insert_count, "   Records:", insert_count, "   Errors:", len (problem))
    msg = "Total SelfPays: " + str (insert_count) + " Records: " + str (insert_count) + " Errors: " + str (len (problem)) + "\n"
    T.insert ('1.0', msg)
    # determine which column contains specific data in 2nd block
    for i in range (1, 15):
        try:
            if re.match ("Referring", ws.cell (block2_start - 1, i).value):
                provider_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Practice", ws.cell (block2_start - 1, i).value):
                practice_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Account", ws.cell (block2_start - 1, i).value):
                CaseNo_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Service", ws.cell (block2_start - 1, i).value):
                svc_date_col = i
                break
        except:
            continue
    insert_count = 0
    for z in range (block2_start, block2_stop + 1):
        selfpay_date = ws.cell (z, svc_date_col).value
        selfpay_date = selfpay_date.strftime ('%Y-%m-%d')
        practice = ws.cell (z, practice_col).value
        provider = ws.cell (z, provider_col).value
        CaseNo = ws.cell (z, CaseNo_col).value
        c.execute ("""INSERT or REPLACE INTO nonpay VALUES(?,?,?,?,?,?,?)""",
                   (my_month, selfpay_date, practice, provider, CaseNo,0,0))
        insert_count = insert_count + 1
    con.commit ()
    #  Associate each NONPAY with the proper Rep
    c.execute ("UPDATE nonpay "
               "SET AcctNo = (SELECT accounts.AcctNo FROM accounts "
               "WHERE accounts.MonthlyName = nonpay.Practice),"
               "RepNo =  (SELECT accounts.RepNo FROM accounts WHERE accounts.MonthlyName = nonpay.Practice)"
               "where nonpay.Month=?", (my_month,))
    con.commit()
    c.execute("select practice from NonPay where acctno is null and month=?", (my_month,))
    problem = c.fetchall()
    print (dbName, "updated:", "Total NonPays:  ", insert_count, "   Records: ", insert_count, "   Errors:",len(problem))
    msg = "Total NonPays: " + str (insert_count) + " Records: " + str (insert_count) + " Errors: " + str (len (problem)) + "\n"
    T.insert ('1.0', msg)

    return

def insert_all_data():
    insert_samples()
    insert_monthly_data()
    Check_for_errors()
    return

def create_csv_file():
    def sel():
        global my_month
        my_month = long_month[(var.get ())]
        a=0
        c.execute('select count(repno) from reps')
        reps_loop =   c.fetchone()[0] + 1
        working_dir = os.getcwd ()
        os.makedirs ('GHData/'+my_month, exist_ok=True)
        os.chdir ('GHData/'+ my_month)
        for k in range(1,reps_loop):
            c.execute ('select Date, Practice, Samples, Rep '
                       'from CSV_Samples '
                       'where RepNo = ? and Month = ?'
                       'order by Date, Practice',(k, my_month))
            data = c.fetchall()
            c.execute ('select Date, Practice, CaseNo '
                       'from CSV_SelfPay '
                       'where RepNo = ? and Month = ? '
                       'order by date, Practice',(k, my_month))
            data2 = c.fetchall()
            c.execute ('select Date, Practice, CaseNo '
                       'from CSV_NonPay '
                       'where RepNo = ? and Month = ? '
                       'order by date, practice', (k, my_month))
            data3 = c.fetchall ()
            if len(data) > 0:
                a=a+1
                lname = data[0][3].split(',')[0]
                Samples = my_month+"-"+lname+".csv"
                with open(Samples,'w',newline='') as f_handle:
                    writer = csv.writer(f_handle)
                    for row in data:
                        writer.writerow(row)
                    writer.writerow("")
                    for row in data2:
                        writer.writerow(row)
                    writer.writerow("")
                    for row in data3:
                        writer.writerow(row)
        print(a,"CSV file(s) created")
        msg = str(a) + " CSV files created\n"
        T.insert('1.0', msg)
        os.chdir(working_dir)

    var = IntVar ()
    long_month = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
                  'November', 'December']
    win2 = Toplevel (root)
    win2.geometry ('150x300+100+100')
    win2.title ('Months')
    rb = [0] * 12
    for x in range (0, 12):
        rb[x] = Radiobutton (win2, text=long_month[x], variable=var, value=x, command=sel).pack (anchor=W)
    frame.pack ()
    win2.mainloop ()
    return

def Monthly_summary():
    def sel():
        global my_month
        my_month = long_month[(var.get())]
        print ("Month Selected --> ", my_month)
        msg = 'Month Selected -->' + my_month
        T.insert('1.0',msg)
        sub_window = Tk ()
        frame = Frame (sub_window)
        frame.pack ()
        sub_window.title ('Summary')
        sub_window.geometry ('900x600')
        T2 = Text (sub_window, height=900, width=600)
        T2.tag_configure ('hdr_fmt', font='courier 14 bold', justify='left')
        T2.tag_configure ('name_fmt', font='courier 11 bold', justify='left')
        T2.tag_configure ('rec_fmt', font='courier 10', justify='center')
        T2.pack ()
        tab = " " * 7
        Grand_Total = 0
        #c.execute ('select rep, Samples, SelfPays, NonPays, Net, "Rep Rate", Date, aRepRate, month '
        #           'from Monthly_Summary where month like ?', (my_month,))
        c.execute ('select rep, Samples, SelfPays, NonPays, Net, "Rep Rate", Date, aRepRate, month '
                   'from Monthly_Summary_All where month like ?', (my_month,))
        data = list (c.fetchall ())
        reps_loop = len (data)

        T2.insert (END, tab)
        T2.insert (END, my_month, "hdr_fmt")
        T2.insert (END, "\n\n\n")
        T2.insert (END, "      Sales Rep         Samples     SP       NP        Net    Rate                 Total\n\n",
                  "name_fmt")
        for p in range (0, reps_loop):
            l = len (data[p][0])
            T2.insert (END, tab)
            T2.insert (END, data[p][0], "rec_fmt")
            T2.insert (END, " " * (20 - l))
            T2.insert (END, "{:>4}".format (data[p][1]), "rec_fmt")
            T2.insert (END, tab)
            T2.insert (END, "{:>4}".format (data[p][2]), "rec_fmt")
            T2.insert (END, tab)
            T2.insert (END, "{:>4}".format (data[p][3]), "rec_fmt")
            T2.insert (END, tab)
            T2.insert (END, "{:>4}".format (data[p][4]), "rec_fmt")
            T2.insert (END, tab)
            if strptime (my_month, '%B').tm_mon >= int (data[p][6][5:7]):
                T2.insert (END, "{:>4}".format (data[p][5]), "rec_fmt")
                T2.insert (END, tab)
                Total = data[p][4] * data[p][5]
                T2.insert (END, "          ${:0.2f}".format (Total), "rec_fmt")
            else:
                T2.insert (END, "{:>4}".format (data[p][7]), "rec_fmt")
                T2.insert (END, tab)
                Total = data[p][4] * data[p][7]
                T2.insert (END, "          ${:0.2f}".format (Total), "rec_fmt")
            T2.insert (END, "\n\n")
            Grand_Total = Grand_Total + Total
            if p == reps_loop - 1:
                T2.insert (END, tab * 60)
                T2.insert (END, ("${:>10,.2f}".format (Grand_Total)), "hdr_fmt")

        # Create an html file of monthly data
        html_string = """
           <!DOCTYPE html>
           <html>
           <body>
           <table id="example" class="display" width="50%">
                   <thead>
                       <tr>
                           <th style="text-align: left;width: 150px;line-height: 2.0;"> Sales Person</th>
                           <th style="text-align: left;width: 100px;"> Samples </th>
                           <th style="text-align: left;width: 100px;"> Self Pays </th>
                           <th style="text-align: left;width: 100px;"> Non Pays </th>
                           <th style="text-align: left;width: 100px;"> Net </th>
                           <th style="text-align: left;width: 100px;"> Rate </th>
                           <th style="text-align: left;width: 100px;"> Total </th>
                       </tr>
                   </thead> 
           """

        row_string = "<tr>"
        for p in range (0, reps_loop):
            row_string += "<td>" + str (data[p][0]) + "</td>"
            row_string += "<td>" + str (data[p][1]) + "</td>"
            row_string += "<td>" + str (data[p][2]) + "</td>"
            row_string += "<td>" + str (data[p][3]) + "</td>"
            row_string += "<td>" + str (data[p][4]) + "</td>"
            if strptime (my_month, '%B').tm_mon >= int (data[p][6][5:7]):
                row_string += "<td>" + str (data[p][5]) + "</td>"
                Total = data[p][4] * data[p][5]
                row_string += "<td>" + "$ {:>10,.2f}".format (Total) + "</td>"
            else:
                row_string += "<td>" + str (data[p][7]) + "</td>"
                Total = data[p][4] * data[p][7]
                row_string += "<td>" + "$ {:>10,.2f}".format (Total) + "</td>"
            row_string += "<tr>"
        html_string += row_string

        file_name = 'GHData' + '/' + my_month + ".html"

        Html_file = open (file_name, "w")
        Html_file.write (html_string)
        Html_file.close ()
        msg = '\nHTML file Created: '+ file_name + "\n"
        print ("\nHTML File Created:", file_name)
        T.insert('1.0',msg)
        mainloop ()
        return

    global my_month
    var = IntVar ()
    long_month = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
                  'November', 'December']
    win2 = Toplevel (root)
    win2.geometry ('150x300+100+100')
    win2.title ('Months')
    rb = [0] * 12
    for x in range (0, 12):
        rb[x] = Radiobutton (win2, text=long_month[x], variable=var, value=x, command=sel).pack (anchor=W)
    frame.pack ()
    win2.mainloop ()
    return

def Month_request():  #not used
    def sel():
        global my_month
        my_month = long_month[(var.get ())]
        return

    var = IntVar ()
    long_month = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
                  'November', 'December']
    win2 = Toplevel (root)
    win2.geometry ('150x300+100+100')
    win2.title ('Months')
    rb = [0] * 12
    for x in range (0, 12):
        rb[x] = Radiobutton (win2, text=long_month[x], variable=var, value=x, command=sel).pack (anchor=W)
    frame.pack ()
    win2.mainloop ()
    return

def Check_for_errors():
    global my_month
    c.execute ("select distinct practice from SAMPLEs where acctno is null and month=?", (my_month,))
    problem = c.fetchall ()
    if problem:
        print ("\n")
        T.insert('1.0',"\n")
        for g in range (0, len (problem)):
            print ("    *** ERROR: Cannot identify Account for SAMPLE record --> ", problem[g])
            msg = "SAMPLES-Unknown Practice-->" + str(problem[g]) + "\n"
            T.insert ('1.0', msg)
        print("\n")

    c.execute ("select distinct practice from selfpay where acctno is null and month=?", (my_month,))
    problem = c.fetchall ()
    if problem:
        for g in range (0, len (problem)):
            print ("    *** ERROR: Cannot identify Account for SELFPAY record --> ", problem[g])
            msg = "SELFPAYS-Unknown Practice-->" + str(problem[g]) + "\n"
            T.insert ('1.0', msg)
        print ("\n")

    c.execute("select distinct practice from NonPay where acctno is null and month=?", (my_month,))
    problem = c.fetchall()
    if problem:
        for g in range (0, len (problem)):
            print ("    *** ERROR: Cannot identify Account for NONPAY record --> ", problem[g])
            msg = "NONPAYS-Unknown Practice-->" + str(problem[g]) + "\n"
            T.insert ('1.0', msg)

def we_b_done():
  #  con.commit()
  #  con.close()
    root.destroy()

if __name__ == '__main__':
    try:                ##Open the DB
        global dbName, con, c
        dbName = cfg["db_name"]
        con = sqlite3.connect(dbName,check_same_thread=False)
        c = con.cursor()
    except:
        msg = "ERROR:  Unable to connect to DB"
        print(msg)
    #create the data directory for future use
    os.makedirs('GHData/',exist_ok=True)
    working_dir = os.getcwd()
    frame = Frame(root)
    frame.pack()
    root.title('Gus Health v10-Aug')
    root.geometry ('500x650+300+100')
    Bn = Button (frame, height=2, width=20, font=('Helvetica', '12'),
                 text="DAILY PROCESSOR", command=start_thread).pack ()
    Bn = Button (frame, height=1, width=60, font=('Helvetica', '4'),
                 text="     ", command=None).pack ()
    Bn = Button(frame,height=2,width=20,font=('Helvetica', '12'),
                 text="Import Data", command=get_file).pack()
    Bn = Button (frame, height=2, width=20, font=('Helvetica', '12'),
                 text="Create CSV File", command=create_csv_file).pack ()
    Bn = Button (frame, height=2, width=20, font=('Helvetica', '12'),
                 text="Monthly Summary", command=Monthly_summary).pack ()
    Bn = Button(frame,height=2,width=20,font=('Helvetica', '12'),
                 text="Quit", command=we_b_done).pack()

    T = Text(root)
    #scrollbar = Scrollbar (T)
    #scrollbar.pack (side=RIGHT)
    T.pack ()

    root.mainloop ()
