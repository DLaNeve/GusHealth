#!/usr/bin/python3
# version 6

from tkinter import filedialog
from tkinter import *
import tkinter as tk
import sqlite3
from openpyxl.utils import get_column_letter
from time import strptime
import openpyxl
import csv
import os
import re

root = Tk()

def get_file():
    def sel():
        global my_month, ws
        my_month = wb.sheetnames[(var.get ())]
        ws = wb[my_month]
        print ("Month Selected --> ",my_month)
        return

    global fname  #Todo move all globals together at top of function?
    fname = filedialog.askopenfilename \
        (initialdir="/", title="Select file",
         filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
    wb = openpyxl.load_workbook (fname, data_only=True)
    print('\nFile Selected ---> ',fname)
    wb.sheetnames
    var = IntVar()
    rb = [0]*20
    for x in range(0,len(wb.sheetnames)):
        rb[x] = Radiobutton (root, text=wb.sheetnames[x], variable=var, value=x, command=sel).pack (anchor=W)
    label = Label (root)
    label.pack ()
    root.mainloop()

    wb.close()
    return

def insert_data():
# Find the first row with sample data
    for first_row in range(1,100):
        if ws.cell(first_row,1).value != None:
            break
# Find the last row with sample data
    for last_row in range(first_row,100):
        if ws.cell(last_row,1).value == None:
            last_row = last_row - 1
            num_rows = last_row - 2
            break
# Find the last column with a DATE
    for last_col in range(first_row, 35):
        if ws.cell(1,last_col).value == None:
            last_col = last_col -1
            col_letter = get_column_letter(last_col)
            break
#Create lists based on the row/col from above
    account = []
    for x in range (first_row, last_row+1):
        account.append (ws.cell (x, 1).value)
    doc = []
    for x in range (first_row, last_row+1):
        doc.append (ws.cell (x, 2).value)
# Insert records into DB only if sample >0
    insert_count = 0
    total_samples = 0
    for y in range (4, last_col+1):
        sample_date = ws.cell (1, y).value.strftime ('%Y-%m-%d')
        samples = []
        for x in range (first_row, last_row+1):
            samples.append (ws.cell (x, y).value)
        for x in range (0,num_rows):  ## Insert rows into DB
            if samples[x] is not None and samples[x] > 0:
                c.execute ("""INSERT or REPLACE INTO Samples VALUES(?,?,?,?,?)""",
                          (my_month, sample_date, account[x], doc[x], samples[x]))
                insert_count=insert_count+1
                total_samples = total_samples + samples[x]
        con.commit()
    print(dbName,"updated:", total_samples,"Samples -", insert_count, "records for",my_month)
    return

def self_pay_loader():
    ##Find the row containing the words "Referring... (block 1)
    col_num = 1
    for block1_start in range (1, 500):
        if ws.cell (block1_start, col_num).value == "Referring Provider Name":
        #if re.match ("Referring", ws.cell (block1_start, col_num).value):
            block1_start = block1_start + 1
            break
    ##Find the last row containing data in block 1.
    col_num = 1
    for block1_stop in range (block1_start, 500):
        if ws.cell (block1_stop, col_num).value is None:
            block1_stop = block1_stop - 1
            break
    ##Find the row containing the word "Referring... again  (block 2).
    col_num = 1
    for block2_start in range (block1_stop, 500):
        if ws.cell (block2_start, col_num).value == "Referring Provider Name":
        #if re.match ("Referring", ws.cell (block2_start, col_num).value):
            block2_start = block2_start + 1
            break
    ##Find the last row containing data in block 2
    col_num = 1
    for block2_stop in range (block2_start, 500):
        if ws.cell (block2_stop, col_num).value is None:
            block2_stop = block2_stop - 1
            break
    # determine which column contains specific data in 1st block
    for i in range(1,15):
        try:
            if re.match("Referring",ws.cell(block1_start-1,i).value):
                provider_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Practice", ws.cell (block1_start-1, i).value):
                practice_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Account", ws.cell (block1_start-1, i).value):
                acctnumber_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Service", ws.cell (block1_start-1, i).value):
                svc_date_col = i
                break
        except:
            continue
    insert_count = 0
    for z in range (block1_start, block1_stop + 1):
        selfpay_date = ws.cell (z,svc_date_col).value.strftime ('%Y-%m-%d')
        practice = ws.cell(z,practice_col).value
        provider = ws.cell(z,provider_col).value
        acctnumber = ws.cell(z,acctnumber_col).value
        c.execute ("""INSERT or REPLACE INTO selfpay VALUES(?,?,?,?,?)""",
                   (my_month, selfpay_date, practice, provider, acctnumber))
        insert_count = insert_count + 1
    con.commit ()
    print(dbName, "updated:",  insert_count, "SelfPays -", insert_count, "records for", my_month)

    # determine which column contains specific data in 2nd block
    for i in range (1, 15):
        try:
            if re.match ("Referring", ws.cell (block2_start - 1, i).value):
                provider_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Practice", ws.cell (block2_start - 1, i).value):
                practice_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Account", ws.cell (block2_start - 1, i).value):
                acctnumber_col = i
                break
        except:
            continue
    for i in range (1, 15):
        try:
            if re.match ("Service", ws.cell (block2_start - 1, i).value):
                svc_date_col = i
                break
        except:
            continue
    insert_count = 0
    for z in range (block2_start, block2_stop + 1):
        selfpay_date = ws.cell (z, svc_date_col).value.strftime ('%Y-%m-%d')
        practice = ws.cell (z, practice_col).value
        provider = ws.cell (z, provider_col).value
        acctnumber = ws.cell (z, acctnumber_col).value
        c.execute ("""INSERT or REPLACE INTO nonpay VALUES(?,?,?,?,?)""",
                   (my_month, selfpay_date, practice, provider, acctnumber))
        insert_count = insert_count + 1
    con.commit ()
    print (dbName, "updated:", insert_count, "Non Pays -", insert_count, "records for", my_month)

    return


def create_csv_file():
    a=0
    #month_number = '{0:0>2}'.format (strptime (my_month.split()[0], '%B').tm_mon)
    #srch_date = "2018-" + str (month_number) + "-%"
    c.execute('select count(repno) from reps')
    reps_loop =   c.fetchone()[0] + 1
    working_dir = os.getcwd ()
    os.makedirs ('GHData/'+my_month, exist_ok=True)
    os.chdir ('GHData/'+ my_month)
    for T in range(1,reps_loop):
        c.execute ('select date, Salesperson, ReportName, Doctor, Samples '
                   'from Samples_CSV '
                   'where RepNo = ? and Month = ?'
                   'order by Date, ReportName, doctor',(T, my_month))
        data = c.fetchall()
        c.execute ('select Date, Practice, Doctor, AcctNo from SelfPay_CSV '
                   'where RepNo = ? and Month = ? '
                   'order by date, doctor',(T, my_month))
        data2 = c.fetchall()
        if len(data) > 0:
            a=a+1
            lname = data[0][1].split(',')[0]
            Samples = my_month+"-"+lname+".csv"
            with open(Samples,'w',newline='') as f_handle:
                writer = csv.writer(f_handle)
                for row in data:
                    writer.writerow(row)
                for row in data2:
                    writer.writerow(row)
    print(a,"CSV file(s) created")
    os.chdir(working_dir)
    return


def Monthly_summary():
    sub_window = Tk ()
    frame = Frame (sub_window)
    frame.pack ()
    sub_window.title ('Summary')
    sub_window.geometry ('800x500')
    T = Text(sub_window, height=800, width=500)
    T.tag_configure ('hdr_fmt', font='courier 14 bold', justify='left')
    T.tag_configure ('name_fmt', font='courier 11 bold', justify = 'left')
    T.tag_configure ('rec_fmt', font='courier 10', justify='center')
    T.pack()
    tab = " " * 7
    Grand_Total = 0
    c.execute ('select salesperson, Samples, SelfPays, Net, Rate, Total from Monthly_Summary where month=?',(my_month,))
    data = list(c.fetchall ())
    reps_loop = len(data)
    T.insert (END, tab)
    T.insert(END,my_month,"hdr_fmt")
    T.insert(END,"\n\n\n")
    T.insert(END,"      Sales Rep         Samples     SP       Net     Rate                Total\n\n","name_fmt")
    for p in range (0,reps_loop):
        l=len(data[p][0])
        T.insert (END, tab)
        T.insert (END, data[p][0],"rec_fmt")
        T.insert (END," "*(20-l))
        T.insert (END,  "{:>4}".format(data[p][1]),"rec_fmt")
        T.insert (END, tab)
        T.insert (END, "{:>4}".format (data[p][2]), "rec_fmt")
        T.insert (END, tab)
        T.insert (END, "{:>4}".format (data[p][3]), "rec_fmt")
        T.insert (END, tab)
        T.insert (END, "${:02d}".format (data[p][4]), "rec_fmt")
        T.insert (END, tab*2)
        T.insert (END, "${:>10,.2f}".format(data[p][5]),"rec_fmt")
        T.insert (END,"\n\n")
        Grand_Total = Grand_Total + data[p][5]
        if p==reps_loop-1:
            T.insert (END, tab*39)
            T.insert (END, ("${:>10,.2f}".format(Grand_Total)),"hdr_fmt")
    mainloop()



def we_b_done():
    con.commit()
    con.close()
    root.destroy()

def main():
    try:                ##Open the DB
        global dbName, con, c
        dbName =  "GusHealth_v6.db"
        con = sqlite3.connect(dbName)
        c = con.cursor()
    except:
        print("ERROR:  Unable to connect to DB")
    frame = Frame(root)
    frame.pack()
    root.title('Gus Health v2')
    root.geometry ('350x450')
    Bn = Button(frame,height=2,width=20,font=('Helvetica', '12'),
                    text="Select File", command=get_file).pack()
    Bn = Button (frame, height=2, width=20, font=('Helvetica', '12'),
                    text="Update Samples ", command=insert_data).pack ()
    Bn = Button(frame,height=2,width=20,font=('Helvetica', '12'),
                    text="Import Self Pays", command=self_pay_loader).pack()
    Bn = Button (frame, height=2, width=20, font=('Helvetica', '12'),
                    text="Create CSV", command=create_csv_file).pack ()
    Bn = Button (frame, height=2, width=20, font=('Helvetica', '12'),
                 text="Monthly Summary", command=Monthly_summary).pack ()
    Bn = Button(frame,height=2,width=20,font=('Helvetica', '12'),
                    text="Quit", command=we_b_done).pack()

    root.mainloop ()

    return

main()

